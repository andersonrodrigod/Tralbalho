import os
from openpyxl import load_workbook, Workbook

# --- 1️⃣ Caminho principal ---
pasta_principal = r"C:\Users\anderson.dossantos\Desktop\dev\Tralbaho\gestar"

# --- 2️⃣ Lista de meses ---
meses = ["dez", "jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out"]

# --- 3️⃣ Nome do arquivo consolidado ---
saida_arquivo = os.path.join(pasta_principal, "resultado_concatenado.xlsx")

# --- 4️⃣ Criar planilha final ---
wb_final = Workbook()
ws_final = wb_final.active
ws_final.title = "Consolidado"

# --- 5️⃣ Controle de linhas ---
linha_atual = 1
contagem_linhas = {}

# --- 6️⃣ Função para encontrar onde começa a tabela ---
def encontrar_inicio_tabela(planilha):
    for i, linha in enumerate(planilha.iter_rows(values_only=True), start=1):
        if linha and any("Filial" in str(cel) or "Unidades" in str(cel) for cel in linha if cel):
            return i
    return None

# --- 7️⃣ Percorrer pastas e arquivos ---
for mes in meses:
    pasta_mes = os.path.join(pasta_principal, mes)

    if os.path.exists(pasta_mes):
        for arquivo in os.listdir(pasta_mes):
            if arquivo.lower() == "data 1.xlsx":
                caminho_arquivo = os.path.join(pasta_mes, arquivo)

                wb = load_workbook(caminho_arquivo)
                ws = wb.active

                linha_inicio = encontrar_inicio_tabela(ws)
                if linha_inicio is None:
                    print(f"❌ {mes.upper()}: Cabeçalho não encontrado.")
                    continue

                # --- Se for o primeiro mês, copiar cabeçalho ---
                if linha_atual == 1:
                    for linha in ws.iter_rows(min_row=linha_inicio, max_row=linha_inicio):
                        for cel in linha:
                            nova_cel = ws_final.cell(row=linha_atual, column=cel.col_idx, value=cel.value)
                            if cel.has_style:
                                nova_cel._style = cel._style
                    linha_atual += 1

                # --- Copiar todas as linhas seguintes ---
                linhas_copiadas = 0
                for linha in ws.iter_rows(min_row=linha_inicio + 1, values_only=False):
                    valores = [cel.value for cel in linha]
                    if all(v is None for v in valores):
                        continue  # pular linhas totalmente vazias
                    for cel in linha:
                        nova_cel = ws_final.cell(row=linha_atual, column=cel.col_idx, value=cel.value)
                        if cel.has_style:
                            nova_cel._style = cel._style
                    linha_atual += 1
                    linhas_copiadas += 1

                contagem_linhas[mes] = linhas_copiadas

    else:
        print(f"⚠️ Pasta {mes} não encontrada.")

# --- 8️⃣ Salvar arquivo final ---
wb_final.save(saida_arquivo)

# --- 9️⃣ Exibir contagem ---
total_linhas = sum(contagem_linhas.values())
print("\n📊 CONTAGEM DE LINHAS POR MÊS:")
for mes, qtd in contagem_linhas.items():
    print(f"{mes.upper():>3} → {qtd} linhas")

print(f"\n🧮 TOTAL GERAL: {total_linhas} linhas")
print(f"\n✅ Arquivo final salvo em:\n{saida_arquivo}")
