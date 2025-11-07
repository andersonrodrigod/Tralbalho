import pandas as pd
import warnings
from openpyxl import load_workbook

warnings.simplefilter("ignore", UserWarning)

# --- 1. Arquivos ---
nao_respondidos = "base_nao_respondidos.xlsx"
base_principal = "Base Colaboradores Enfermagem.xlsx"
saida = "Base Colaboradores Enfermagem_com_marcacao.xlsx"

# --- 2. Ler o arquivo base_nao_respondidos ---
df_nao = pd.read_excel(nao_respondidos)

# --- 3. Filtrar apenas os que têm STATUS = REALIZADO ---
df_realizados = df_nao[df_nao["STATUS"].str.strip().str.upper() == "REALIZADO"]

# --- 4. Extrair os nomes da coluna Colaborador ---
nomes_realizados = df_realizados["Colaborador"].dropna().str.strip().str.upper().tolist()

print(f"🔍 Total de 'REALIZADO' encontrados: {len(nomes_realizados)}")

# --- 5. Ler a aba 'Base' do arquivo principal (linha 2 é o cabeçalho real) ---
df_base = pd.read_excel(base_principal, sheet_name="Base", header=1)

# --- 6. Adicionar coluna auxiliar com nomes padronizados ---
df_base["Colab_upper"] = df_base["Colaborador"].astype(str).str.strip().str.upper()

# --- 7. Carregar o arquivo original com openpyxl (para editar diretamente a aba) ---
wb = load_workbook(base_principal)
ws = wb["Base"]

# --- 8. Verificar em qual índice está a coluna AE (caso varie no futuro) ---
ultima_coluna = ws.max_column
coluna_ae = ultima_coluna  # AE é a última

# --- 9. Marcar “encontrado manualmente” onde houver correspondência ---
cont = 0
for i, nome in enumerate(df_base["Colab_upper"], start=3):  # começa da linha 3, pois header=1
    if nome in nomes_realizados:
        ws.cell(row=i, column=coluna_ae, value="encontrado manualmente")
        cont += 1

print(f"✅ Total marcados como 'encontrado manualmente': {cont}")

# --- 10. Salvar em novo arquivo ---
wb.save(saida)
print(f"📁 Arquivo salvo com sucesso: {saida}")
