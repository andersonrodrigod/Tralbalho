import pandas as pd
from openpyxl import load_workbook

# --- Caminhos dos arquivos ---
arquivo_destino = "Planilha agosto atualizada 16.10.xlsx"        # arquivo principal (com fórmulas)
arquivo_origem = "planilha_tratada_status.xlsx"                 # arquivo com os novos dados
aba_alvo = "status"                                             # aba a atualizar

# --- 1️⃣ Ler a planilha de origem ---
print(f"📥 Lendo dados da planilha de origem: {arquivo_origem}")
df_origem = pd.read_excel(arquivo_origem)

# --- 2️⃣ Abrir o arquivo principal com openpyxl ---
print(f"🔍 Abrindo planilha principal: {arquivo_destino}")
wb = load_workbook(arquivo_destino)

# Verifica se a aba existe
if aba_alvo not in wb.sheetnames:
    print(f"❌ A aba '{aba_alvo}' não foi encontrada!")
else:
    ws = wb[aba_alvo]

    # --- 3️⃣ Limpar a aba mantendo o cabeçalho original ---
    print(f"🧹 Limpando conteúdo da aba '{aba_alvo}' (mantendo cabeçalho)...")
    colunas = [cell.value for cell in ws[1]]  # cabeçalho da primeira linha

    # Remove todas as linhas, exceto a primeira (cabeçalho)
    ws.delete_rows(2, ws.max_row)

    # --- 4️⃣ Repreencher com os dados da planilha de origem ---
    # Pega apenas as colunas em comum
    colunas_comuns = [col for col in df_origem.columns if col in colunas]

    if not colunas_comuns:
        print("⚠️ Nenhuma coluna em comum entre os arquivos. Nenhum dado será inserido.")
    else:
        print(f"📊 Inserindo dados nas colunas: {colunas_comuns}")

        # Reorganiza o DataFrame para seguir a ordem das colunas da aba destino
        df_final = df_origem[colunas_comuns]

        # Adiciona linha por linha (mantém fórmulas das outras abas)
        for row in df_final.itertuples(index=False, name=None):
            ws.append(row)

        # --- 5️⃣ Salvar as alterações ---
        wb.save(arquivo_destino)
        print(f"✅ Aba '{aba_alvo}' atualizada com sucesso!")
