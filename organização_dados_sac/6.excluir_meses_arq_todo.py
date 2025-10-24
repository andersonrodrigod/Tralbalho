import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

warnings.simplefilter("ignore", UserWarning)

# Arquivo original
arquivo = "Planilha Julho 20.10.xlsx"

# Lista de abas que devem ser filtradas
abas = [
    "p1", "comen p1",
    "p2", "comen p2",
    "p3", "comen p3",
    "p4", "comen p4",
    "p5", "comen p5",
    "P6", "comen p6",
]

# Carrega o arquivo Excel existente (mantém fórmulas das outras abas)
wb = openpyxl.load_workbook(arquivo)

for aba in abas:
    if aba not in wb.sheetnames:
        print(f"⚠️ Aba '{aba}' não encontrada. Pulando...")
        continue

    # Lê a aba com pandas
    df = pd.read_excel(arquivo, sheet_name=aba)

    # Garante que existe a coluna "Nome"
    if "Nome" not in df.columns:
        print(f"⚠️ Aba '{aba}' não tem coluna 'Nome'. Pulando...")
        continue

    # --- Filtro e substituição ---
    df_filtrado = df[df["Nome"].astype(str).str.contains("julho", case=False, na=False)]
    df_filtrado["Nome"] = df_filtrado["Nome"].astype(str).str.replace("-Julho", "_Julho", case=False)

    # --- Substitui conteúdo da aba ---
    ws = wb[aba]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df_filtrado, index=False, header=True):
        ws.append(r)

    print(f"✅ Aba '{aba}' atualizada ({len(df_filtrado)} linhas mantidas).")

# Salva mantendo o restante intacto (fórmulas, estilos, etc.)
wb.save(arquivo)
print("\n🎯 Atualização concluída com sucesso! As demais abas foram preservadas.")
