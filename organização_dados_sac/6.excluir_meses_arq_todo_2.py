import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

warnings.simplefilter("ignore", UserWarning)

# 📘 Arquivo principal
arquivo = "Planilha agosto atualizada 16.10.xlsx"

# 🧾 Abas a processar (exceto a 'status')
abas = [
    "p1", "comen p1",
    "p2", "comen p2",
    "p3", "comen p3",
    "p4", "comen p4",
    "p5", "comen p5",
    "P6", "comen p6",
    "status"  # incluímos aqui também
]

# ⚙️ Carrega o Excel mantendo fórmulas e formatações
wb = openpyxl.load_workbook(arquivo)

for aba in abas:
    if aba not in wb.sheetnames:
        print(f"⚠️ Aba '{aba}' não encontrada. Pulando...")
        continue

    # Lê a aba atual
    df = pd.read_excel(arquivo, sheet_name=aba)

    # Define o nome da coluna a ser usada conforme a aba
    coluna_chave = "Contato" if aba.lower() == "status" else "Nome"

    # Verifica se a coluna existe
    if coluna_chave not in df.columns:
        print(f"⚠️ Aba '{aba}' não tem coluna '{coluna_chave}'. Pulando...")
        continue

    # 🔍 Filtro e substituição
    df_filtrado = df[df[coluna_chave].astype(str).str.contains("agosto", case=False, na=False)]
    df_filtrado[coluna_chave] = df_filtrado[coluna_chave].astype(str).str.replace("-Agosto", "_Agosto", case=False)

    # ✏️ Substitui o conteúdo da aba (mantendo cabeçalho)
    ws = wb[aba]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df_filtrado, index=False, header=True):
        ws.append(r)

    print(f"✅ Aba '{aba}' atualizada ({len(df_filtrado)} linhas mantidas).")

# 💾 Salva o arquivo final
wb.save(arquivo)
print("\n🎯 Atualização concluída com sucesso! Todas as demais abas e fórmulas foram preservadas.")
