from openpyxl import load_workbook

caminho_arquivo = 'detalhamento_de_pesquisa_urgencia.xlsx'
meses_para_remover = ['Julho de 2025', 'Agosto de 2025']

wb = load_workbook(caminho_arquivo)

for aba in wb.sheetnames:
    print(f"🔄 Processando aba: {aba}")
    planilha = wb[aba]
    linhas_para_apagar = []

    for linha in planilha.iter_rows():
        for celula in linha:
            if celula.value and any(mes in str(celula.value) for mes in meses_para_remover):
                linhas_para_apagar.append(celula.row)
                break

    for linha in sorted(set(linhas_para_apagar), reverse=True):
        planilha.delete_rows(linha)

    print(f"✅ {len(linhas_para_apagar)} linhas removidas da aba '{aba}'")

wb.save('detalhamento_de_pesquisa_urgencia_limpo.xlsx')
print("💾 Arquivo salvo com sucesso!")
