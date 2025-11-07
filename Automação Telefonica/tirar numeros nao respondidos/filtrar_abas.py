from openpyxl import load_workbook, Workbook

# Caminho da planilha original
caminho_arquivo_original = 'Planilha JUNHO 01.10.xlsx'

# Nome da nova planilha
caminho_nova_planilha = 'Planilha JUNHO 01.10 Filtrada.xlsx'

# Carrega a planilha original
wb_original = load_workbook(caminho_arquivo_original)

# Cria uma nova planilha
wb_nova = Workbook()

# Remove a aba padrão criada automaticamente
wb_nova.remove(wb_nova.active)

# Lista de abas que você quer manter
abas_desejadas = ['BASE', 'status']

for aba in abas_desejadas:
    if aba in wb_original.sheetnames:
        folha = wb_original[aba]
        nova_folha = wb_nova.create_sheet(title=aba)
        for linha in folha.iter_rows(values_only=True):
            nova_folha.append(linha)

# Salva a nova planilha
wb_nova.save(caminho_nova_planilha)
print("Nova planilha criada com sucesso!")
