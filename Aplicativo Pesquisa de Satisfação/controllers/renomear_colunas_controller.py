# controllers/renomear_colunas_controller.py
import pandas as pd
from openpyxl import load_workbook
import os

class RenomearColunasController:
    def __init__(self):
        self.arquivo_selecionado = None
        self.workbook = None
        self.abas_originais = []
        self.colunas_por_aba = {}

    def carregar_arquivo(self, arquivo_path):
        """Carrega o arquivo Excel e extrai informações das abas e colunas"""
        try:
            self.arquivo_selecionado = arquivo_path
            
            # Carregar com openpyxl para modificações
            self.workbook = load_workbook(arquivo_path)
            self.abas_originais = self.workbook.sheetnames.copy()

            print("Abas encontradas no arquivo:")
            for aba in self.workbook.sheetnames:
                print(f"'{aba}'")
            
            # Obter colunas de cada aba usando pandas
            self.colunas_por_aba = {}
            excel_file = pd.ExcelFile(arquivo_path)
            
            for aba in self.abas_originais:
                try:
                    # Ler apenas o cabeçalho para obter as colunas
                    df = pd.read_excel(arquivo_path, sheet_name=aba, nrows=0)
                    self.colunas_por_aba[aba] = list(df.columns)
                    print(f"Colunas na aba '{aba}': {self.colunas_por_aba[aba]}")
                except Exception as e:
                    print(f"Erro ao ler colunas da aba '{aba}': {e}")
                    self.colunas_por_aba[aba] = []
            
            return {
                'sucesso': True,
                'abas': self.abas_originais,
                'colunas_por_aba': self.colunas_por_aba,
                'total_abas': len(self.abas_originais)
            }
            
        except Exception as e:
            return {
                'sucesso': False,
                'erro': str(e)
            }

    def renomear_coluna(self, aba, coluna_antiga, coluna_nova):
        """Renomeia uma coluna específica em uma aba"""
        try:
            # Verificar se a aba existe
            if aba not in self.colunas_por_aba:
                return {'sucesso': False, 'erro': f'A aba "{aba}" não foi encontrada!'}
            
            # Verificar se a coluna antiga existe
            if coluna_antiga not in self.colunas_por_aba[aba]:
                return {'sucesso': False, 'erro': f'A coluna "{coluna_antiga}" não foi encontrada na aba "{aba}"!'}
            
            # Verificar se a coluna nova já existe (e não é a mesma)
            if coluna_nova in self.colunas_por_aba[aba] and coluna_nova != coluna_antiga:
                return {'sucesso': False, 'erro': f'A coluna "{coluna_nova}" já existe na aba "{aba}"!'}
            
            # Atualizar no dicionário de colunas
            index = self.colunas_por_aba[aba].index(coluna_antiga)
            self.colunas_por_aba[aba][index] = coluna_nova
            
            print(f"Coluna renomeada: {aba}.{coluna_antiga} → {coluna_nova}")
            
            return {'sucesso': True}
            
        except Exception as e:
            return {'sucesso': False, 'erro': str(e)}

    def salvar_arquivo(self, caminho, salvar_como_novo=True):
        """Salva o arquivo com as modificações das colunas"""
        try:
            # Para aplicar renomeações de colunas, precisamos processar cada aba
            dados_processados = {}
            
            for aba in self.abas_originais:
                # Ler os dados da aba original
                df = pd.read_excel(self.arquivo_selecionado, sheet_name=aba)
                
                # Aplicar renomeações de colunas
                colunas_atuais = list(df.columns)
                colunas_renomeadas = self.colunas_por_aba.get(aba, colunas_atuais)
                
                # Criar mapeamento de renomeação
                mapeamento = {}
                for atual, novo in zip(colunas_atuais, colunas_renomeadas):
                    if atual != novo:
                        mapeamento[atual] = novo
                
                if mapeamento:
                    print(f"Aplicando renomeações na aba '{aba}': {mapeamento}")
                    df = df.rename(columns=mapeamento)
                
                dados_processados[aba] = df
            
            # Salvar o arquivo processado
            with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
                for aba, df in dados_processados.items():
                    df.to_excel(writer, sheet_name=aba, index=False)
            
            return {'sucesso': True}
            
        except Exception as e:
            return {'sucesso': False, 'erro': str(e)}