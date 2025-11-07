# controllers/renomear_abas_controller.py
import pandas as pd
from openpyxl import load_workbook
import os

class RenomearAbasController:
    def __init__(self):
        self.arquivo_selecionado = None
        self.workbook = None
        self.abas_originais = []

    def carregar_arquivo(self, arquivo_path):
        """Carrega o arquivo Excel e extrai informações das abas"""
        try:
            self.arquivo_selecionado = arquivo_path
            
            # Carregar com openpyxl para modificações
            self.workbook = load_workbook(arquivo_path)
            self.abas_originais = self.workbook.sheetnames.copy()

            print("Abas encontradas no arquivo:")
            for aba in self.workbook.sheetnames:
                print(f"'{aba}'")
            
            return {
                'sucesso': True,
                'abas': self.abas_originais,
                'total_abas': len(self.abas_originais)
            }
            
        except Exception as e:
            return {
                'sucesso': False,
                'erro': str(e)
            }

    def renomear_aba(self, nome_antigo, nome_novo):
        """Renomeia uma aba específica"""
        try:
            # Verifica se a aba foi renomeada anteriormente
            nome_real = None
            for aba in self.workbook.sheetnames:
                if aba == nome_antigo or aba.strip().lower() == nome_antigo.strip().lower():
                    nome_real = aba
                    break

            if not nome_real:
                # Caso tenha sido renomeada anteriormente, tentar localizar pelo histórico
                for aba in self.abas_originais:
                    if aba == nome_antigo:
                        nome_real = aba
                        break

            if not nome_real:
                return {'sucesso': False, 'erro': f'A aba "{nome_antigo}" não foi encontrada!'}

            print(f"Tentando renomear: '{nome_real}' → '{nome_novo}'")

            # Verificar se o novo nome já existe
            if nome_novo in self.workbook.sheetnames and nome_novo != nome_real:
                return {'sucesso': False, 'erro': f'A aba "{nome_novo}" já existe!'}

            # Renomear a aba
            self.workbook[nome_real].title = nome_novo

            # Atualizar as listas internas
            if nome_real in self.abas_originais:
                idx = self.abas_originais.index(nome_real)
                self.abas_originais[idx] = nome_novo
            elif nome_novo not in self.abas_originais:
                self.abas_originais.append(nome_novo)

            return {'sucesso': True}

        except Exception as e:
            return {'sucesso': False, 'erro': str(e)}

    def salvar_arquivo(self, caminho, salvar_como_novo=True):
        """Salva o arquivo com as modificações"""
        try:
            if salvar_como_novo:
                # Para novo arquivo, salvar diretamente o workbook
                self.workbook.save(caminho)
            else:
                # Para sobrescrever, salvar no mesmo caminho
                self.workbook.save(caminho)
            
            return {'sucesso': True}
            
        except Exception as e:
            return {'sucesso': False, 'erro': str(e)}