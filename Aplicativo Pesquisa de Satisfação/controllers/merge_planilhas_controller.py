# controllers/merge_arquivos_controller.py
import pandas as pd
from openpyxl import load_workbook
import os

class MergePlanilhasController:
    def processar_merge(self, arquivo_principal, arquivo_merge, config):
        """
        Processa o merge entre dois arquivos Excel preservando fórmulas
        Usa a mesma lógica do código fornecido pelo usuário
        """
        try:
            # Ler ambos os arquivos
            principal_excel = pd.ExcelFile(arquivo_principal)
            merge_excel = pd.ExcelFile(arquivo_merge)
            
            # Obter listas de abas
            abas_principal = principal_excel.sheet_names
            abas_merge = merge_excel.sheet_names
            
            # Encontrar abas em comum
            abas_comuns = set(abas_principal) & set(abas_merge)
            abas_apenas_principal = set(abas_principal) - set(abas_merge)
            
            # Abrir o arquivo principal com openpyxl para preservar fórmulas
            wb_principal = load_workbook(arquivo_principal)
            
            # Processar cada aba em comum
            for aba in abas_comuns:
                print(f"🔄 Processando aba: {aba}")
                
                # Ler dados da aba do arquivo merge
                df_merge = pd.read_excel(arquivo_merge, sheet_name=aba)
                
                # Obter a worksheet correspondente no arquivo principal
                ws_principal = wb_principal[aba]
                
                # Limpar a aba mantendo o cabeçalho original
                colunas = [cell.value for cell in ws_principal[1]]  # cabeçalho da primeira linha
                
                # Remove todas as linhas, exceto a primeira (cabeçalho)
                if ws_principal.max_row > 1:
                    ws_principal.delete_rows(2, ws_principal.max_row - 1)
                
                # Pegar apenas as colunas em comum entre os dados do merge e a estrutura da planilha
                colunas_comuns = [col for col in df_merge.columns if col in colunas]
                
                if not colunas_comuns:
                    print(f"⚠️ Nenhuma coluna em comum na aba '{aba}'. Nenhum dado será inserido.")
                    continue
                
                print(f"📊 Inserindo dados nas colunas: {colunas_comuns}")
                
                # Reorganiza o DataFrame para seguir a ordem das colunas da aba destino
                df_final = df_merge[colunas_comuns]
                
                # Adiciona linha por linha (mantém fórmulas das outras abas)
                for row in df_final.itertuples(index=False, name=None):
                    ws_principal.append(row)
            
            # As abas apenas do arquivo principal já estão preservadas com fórmulas intactas
            # pois estamos trabalhando diretamente no workbook original
            
            return wb_principal, abas_comuns, abas_apenas_principal
            
        except Exception as e:
            raise Exception(f"Erro ao processar merge: {str(e)}")

    def verificar_compatibilidade_arquivos(self, arquivo_principal, arquivo_merge):
        """
        Verifica se os arquivos são compatíveis para merge
        """
        try:
            principal_excel = pd.ExcelFile(arquivo_principal)
            merge_excel = pd.ExcelFile(arquivo_merge)
            
            abas_principal = principal_excel.sheet_names
            abas_merge = merge_excel.sheet_names
            
            abas_comuns = set(abas_principal) & set(abas_merge)
            
            # Verificar estrutura das colunas para cada aba em comum
            detalhes_abas = {}
            for aba in abas_comuns:
                df_principal = pd.read_excel(arquivo_principal, sheet_name=aba, nrows=1)  # Só cabeçalho
                df_merge = pd.read_excel(arquivo_merge, sheet_name=aba, nrows=1)  # Só cabeçalho
                
                colunas_principal = set(df_principal.columns)
                colunas_merge = set(df_merge.columns)
                colunas_comuns = colunas_principal & colunas_merge
                
                detalhes_abas[aba] = {
                    'colunas_principal': list(colunas_principal),
                    'colunas_merge': list(colunas_merge),
                    'colunas_comuns': list(colunas_comuns),
                    'total_colunas_principal': len(colunas_principal),
                    'total_colunas_merge': len(colunas_merge),
                    'total_colunas_comuns': len(colunas_comuns)
                }
            
            return {
                'abas_principal': abas_principal,
                'abas_merge': abas_merge,
                'abas_comuns': list(abas_comuns),
                'total_abas_principal': len(abas_principal),
                'total_abas_merge': len(abas_merge),
                'total_abas_comuns': len(abas_comuns),
                'detalhes_abas': detalhes_abas
            }
            
        except Exception as e:
            raise Exception(f"Erro ao verificar compatibilidade: {str(e)}")

    def salvar_workbook(self, workbook, caminho):
        """
        Salva o workbook processado
        """
        try:
            workbook.save(caminho)
            return True
        except Exception as e:
            raise Exception(f"Erro ao salvar arquivo: {str(e)}")