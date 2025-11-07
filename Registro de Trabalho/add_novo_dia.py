import pandas as pd
from datetime import date
import unicodedata
import os
from openpyxl import load_workbook

# --- Caminhos dos arquivos ---
resumo_antigo = "RESUMO_DIARIO_OPERADORES.xlsx"  # resumo acumulado
arquivo_base = "Planilha Julho 23.10.xlsx"       # base atual (nova)
saida = "RESUMO_DIARIO_OPERADORES_ATUALIZADO.xlsx"

# --- Funções de limpeza ---
def normalizar_texto(txt):
    txt = str(txt).strip().lower()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("utf-8")
    if "whats" in txt:
        return "whats app"
    elif "liga" in txt:
        return "ligacao"
    else:
        return txt

def normalizar_operador(nome):
    nome = str(nome).strip().lower()
    nome = unicodedata.normalize("NFKD", nome).encode("ascii", "ignore").decode("utf-8")
    return nome.title()

# --- Ler base atual ---
df = pd.read_excel(arquivo_base, sheet_name="BASE")
df.columns = df.columns.str.lower().str.strip()
data_atual = date.today()

# --- Limpeza e normalização ---
df = df.dropna(subset=["operador", "tipo de contato"])
df["operador"] = df["operador"].apply(normalizar_operador)
df["tipo de contato"] = df["tipo de contato"].apply(normalizar_texto)
df["data"] = data_atual

# --- Contagem por operador e tipo ---
tipos = ["whats app", "ligacao"]
todos_operadores = sorted(df["operador"].unique())
combinacoes = pd.MultiIndex.from_product(
    [[data_atual], todos_operadores, tipos],
    names=["data", "operador", "tipo de contato"]
)

resumo_novo = (
    df.groupby(["data", "operador", "tipo de contato"])
    .size()
    .reindex(combinacoes, fill_value=0)
    .reset_index(name="quantidade")
)

# --- Pivot para formato final (igual antes) ---
tabela_whats = resumo_novo[resumo_novo["tipo de contato"] == "whats app"].pivot(
    index="data", columns="operador", values="quantidade"
).fillna(0)

tabela_ligacao = resumo_novo[resumo_novo["tipo de contato"] == "ligacao"].pivot(
    index="data", columns="operador", values="quantidade"
).fillna(0)

tabela_dia = pd.concat([tabela_whats, tabela_ligacao], axis=1, keys=["WHATSAPP", "LIGAÇÃO"])
tabela_dia.index.name = "data"

# --- Verificar se já existe um resumo anterior ---
if os.path.exists(resumo_antigo):
    print("📂 Lendo resumo anterior...")
    resumo_anterior = pd.read_excel(resumo_antigo, header=[0, 1], index_col=0)

    # --- Garantir que todos os operadores antigos e novos estejam presentes ---
    todos_operadores = sorted(set(resumo_anterior.columns.get_level_values(1)) | set(tabela_dia.columns.get_level_values(1)))

    # --- Reindexar os dois para incluir todos os operadores ---
    resumo_anterior = resumo_anterior.reindex(columns=pd.MultiIndex.from_product([["WHATSAPP", "LIGAÇÃO"], todos_operadores]), fill_value=0)
    tabela_dia = tabela_dia.reindex(columns=pd.MultiIndex.from_product([["WHATSAPP", "LIGAÇÃO"], todos_operadores]), fill_value=0)

    # --- Calcular o incremento do dia em relação ao total acumulado ---
    total_acumulado = resumo_anterior.apply(pd.to_numeric, errors="coerce").fillna(0).sum()

    tabela_dia = tabela_dia - total_acumulado

    tabela_dia = tabela_dia.clip(lower=0)

    # --- Concatenar os dados (mantendo histórico) ---
    resumo_final = pd.concat([resumo_anterior, tabela_dia])
else:
    print("🆕 Nenhum resumo anterior encontrado — criando novo.")
    resumo_final = tabela_dia



resumo_final.index = pd.to_datetime(resumo_final.index).date

# --- Salvar arquivo atualizado ---
resumo_final.to_excel(saida, merge_cells=True)
print(f"✅ Resumo atualizado salvo como: '{saida}'")

# --- Ajustes no arquivo salvo (linha 3 e célula A2) ---
wb = load_workbook(saida)
ws = wb.active

# Excluir linha 3
ws.delete_rows(3)

# Colocar o nome "data" na célula A2
ws["A2"] = "data"

wb.save(saida)
wb.close()

print("🧹 Linha 3 removida e célula A2 nomeada como 'data'.")
