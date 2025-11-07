import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

# Carregar os dados do arquivo Excel
df = pd.read_excel('dados_organizados.xlsx', engine='openpyxl')

# Renomear colunas para facilitar
df.columns = ['Responsavel', 'Pergunta', 'Resposta']

# Agrupar por Responsável e Pergunta e calcular o desvio padrão
resultado = df.groupby(['Responsavel', 'Pergunta'])['Resposta'].std().reset_index()

# Arredondar o desvio padrão para uma casa decimal
resultado['Desvio_Padrao'] = resultado['Resposta'].round(1)
resultado.drop(columns='Resposta', inplace=True)

# Criar um novo arquivo Excel
wb = Workbook()
wb.remove(wb.active)  # Remover a planilha padrão

# Criar uma planilha para cada responsável
for i, responsavel in enumerate(resultado['Responsavel'].unique()):
    ws = wb.create_sheet(title=responsavel[:31])  # Nome da aba limitado a 31 caracteres
    df_responsavel = resultado[resultado['Responsavel'] == responsavel].copy()
    df_responsavel.drop(columns='Responsavel', inplace=True)

    # Adicionar os dados à planilha
    for r_idx, row in enumerate(dataframe_to_rows(df_responsavel, index=False, header=True), 1):
        ws.append(row)

    # Criar tabela com nome único
    table_name = f"Tabela_{i}_{responsavel.replace(' ', '_').replace('.', '').replace('-', '').replace('ã', 'a').replace('é', 'e')}"
    tab = Table(displayName=table_name[:30], ref=f"A1:B{len(df_responsavel)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Aplicar formato numérico com uma casa decimal
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row):
        for cell in row:
            cell.number_format = '0.0'

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# Salvar o arquivo
wb.save('desvios_padrao_por_responsavel.xlsx')