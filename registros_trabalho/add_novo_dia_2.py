import pandas as pd
from datetime import date
import unicodedata
import os
from openpyxl import load_workbook

# --- Caminhos dos arquivos ---
resumo_antigo = "RESUMO_DIARIO_OPERADORES.xlsx"  # resumo acumulado
arquivo_base = "Planilha Julho 23.10.xlsx"       # base atual (nova)
saida = "RESUMO_DIARIO_OPERADORES_ATUALIZADO.xlsx"

# --- Funções de limpeza ---
def normalizar_texto(txt):
    txt = str(txt).strip().lower()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("utf-8")
    if "whats" in txt:
        return "whats app"
    elif "liga" in txt:
        return "ligacao"
    else:
        return txt

def normalizar_operador(nome):
    nome = str(nome).strip().lower()
    nome = unicodedata.normalize("NFKD", nome).encode("ascii", "ignore").decode("utf-8")
    return nome.title()

# --- Ler base atual ---
df = pd.read_excel(arquivo_base, sheet_name="BASE")
df.columns = df.columns.str.lower().str.strip()
data_atual = date.today()

# --- Limpeza e normalização ---
df = df.dropna(subset=["operador", "tipo de contato"])
df["operador"] = df["operador"].apply(normalizar_operador)
df["tipo de contato"] = df["tipo de contato"].apply(normalizar_texto)
df["data"] = data_atual

# --- Contagem por operador e tipo ---
tipos = ["whats app", "ligacao"]
todos_operadores = sorted(df["operador"].unique())
combinacoes = pd.MultiIndex.from_product(
    [[data_atual], todos_operadores, tipos],
    names=["data", "operador", "tipo de contato"]
)

resumo_novo = (
    df.groupby(["data", "operador", "tipo de contato"])
    .size()
    .reindex(combinacoes, fill_value=0)
    .reset_index(name="quantidade")
)

# --- Pivot para formato final ---
tabela_whats = resumo_novo[resumo_novo["tipo de contato"] == "whats app"].pivot(
    index="data", columns="operador", values="quantidade"
).fillna(0)

tabela_ligacao = resumo_novo[resumo_novo["tipo de contato"] == "ligacao"].pivot(
    index="data", columns="operador", values="quantidade"
).fillna(0)

tabela_dia = pd.concat(
    [tabela_whats, tabela_ligacao], axis=1, keys=["WHATSAPP", "LIGAÇÃO"]
)
tabela_dia.index.name = "data"

# --- Verificar se já existe um resumo anterior ---
if os.path.exists(resumo_antigo):
    print("📂 Lendo resumo anterior...")
    resumo_anterior = pd.read_excel(resumo_antigo, header=[0, 1], index_col=0)

    # 🔍 Remover linha TOTAL GERAL se existir
    if "TOTAL" in resumo_anterior.index:
        print("🧹 Removendo linha 'TOTAL' do resumo anterior...")
        resumo_anterior = resumo_anterior.drop("TOTAL", axis=0, errors="ignore")

    # --- Garantir que todos os operadores antigos e novos estejam presentes ---
    todos_operadores = sorted(
        set(resumo_anterior.columns.get_level_values(1))
        | set(tabela_dia.columns.get_level_values(1))
    )

    # --- Reindexar os dois para incluir todos os operadores ---
    resumo_anterior = resumo_anterior.reindex(
        columns=pd.MultiIndex.from_product(
            [["WHATSAPP", "LIGAÇÃO"], todos_operadores]
        ),
        fill_value=0,
    )
    tabela_dia = tabela_dia.reindex(
        columns=pd.MultiIndex.from_product(
            [["WHATSAPP", "LIGAÇÃO"], todos_operadores]
        ),
        fill_value=0,
    )

    # --- Calcular o incremento do dia em relação ao total acumulado ---
    total_acumulado = resumo_anterior.apply(pd.to_numeric, errors="coerce").fillna(0).sum()
    tabela_dia = tabela_dia - total_acumulado
    tabela_dia = tabela_dia.clip(lower=0)

    resumo_final = pd.concat([resumo_anterior, tabela_dia])
else:
    print("🆕 Nenhum resumo anterior encontrado — criando novo.")
    resumo_final = tabela_dia

# --- Corrigir formato da data (sem hora) ---
resumo_final.index = pd.to_datetime(resumo_final.index).strftime("%Y-%m-%d")

# --- Salvar sem linha em branco extra ---
with pd.ExcelWriter(saida, engine="openpyxl") as writer:
    resumo_final.to_excel(writer, merge_cells=True)
print(f"✅ Resumo atualizado salvo como: '{saida}'")

# --- Ajustes no arquivo salvo (remover linha 3 e renomear A2) ---
wb = load_workbook(saida)
ws = wb.active

linha3_valores = [cell.value for cell in ws[3]]
if all(v is None for v in linha3_valores):
    ws.delete_rows(3)

ws["A2"] = "data"

wb.save(saida)
wb.close()

print("🧹 Linha 3 vazia removida e cabeçalho 'data' corrigido.")

# ======================================================
# 📊 ETAPA EXTRA — SOMAR SOMENTE CÉLULAS NUMÉRICAS POR OPERADOR
# ======================================================

print("📈 Calculando totais numéricos por operador...")

df_final = pd.read_excel(saida, header=[0, 1], index_col=0)

# Somar apenas valores numéricos
total_por_operador = df_final.apply(
    lambda col: pd.to_numeric(col, errors="coerce").sum(), axis=0
)

# Adicionar linha TOTAL GERAL
linha_total = pd.DataFrame(total_por_operador).T
linha_total.index = ["TOTAL"]

# Concatenar ao final do DataFrame original
df_com_total = pd.concat([df_final, linha_total])

# Garantir que o índice (data) não venha com hora
df_com_total.index = df_com_total.index.map(
    lambda x: x if x == "TOTAL" else str(x).split(" ")[0]
)

# Salvar novamente sem criar linha em branco
with pd.ExcelWriter(saida, engine="openpyxl") as writer:
    df_com_total.to_excel(writer, merge_cells=True)

print("✅ Linha TOTAL GERAL adicionada e tratada para próxima execução.")
